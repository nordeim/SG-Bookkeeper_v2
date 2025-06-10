# File: app/ui/reports/report_engine.py
from typing import Dict, Any, Literal, List, Optional, TYPE_CHECKING, cast 
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepInFrame
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle 
from reportlab.lib import colors 
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch, cm 
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from io import BytesIO
import openpyxl 
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill 
from openpyxl.utils import get_column_letter 
from decimal import Decimal, InvalidOperation
from datetime import date

from app.utils.pydantic_models import GSTReturnData, GSTTransactionLineDetail

if TYPE_CHECKING:
    from app.core.application_core import ApplicationCore 

class ReportEngine:
    def __init__(self, app_core: "ApplicationCore"): 
        self.app_core = app_core
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()

    def _setup_custom_styles(self):
        self.styles.add(ParagraphStyle(name='ReportTitle', parent=self.styles['h1'], alignment=TA_CENTER, spaceAfter=0.2*inch))
        self.styles.add(ParagraphStyle(name='ReportSubTitle', parent=self.styles['h2'], alignment=TA_CENTER, fontSize=12, spaceAfter=0.2*inch))
        self.styles.add(ParagraphStyle(name='SectionHeader', parent=self.styles['h3'], fontSize=11, spaceBefore=0.2*inch, spaceAfter=0.1*inch, textColor=colors.HexColor("#2F5496")))
        self.styles.add(ParagraphStyle(name='AccountName', parent=self.styles['Normal'], leftIndent=0.2*inch))
        self.styles.add(ParagraphStyle(name='AccountNameBold', parent=self.styles['AccountName'], fontName='Helvetica-Bold'))
        self.styles.add(ParagraphStyle(name='Amount', parent=self.styles['Normal'], alignment=TA_RIGHT))
        self.styles.add(ParagraphStyle(name='AmountBold', parent=self.styles['Amount'], fontName='Helvetica-Bold'))
        self.styles.add(ParagraphStyle(name='TableHeader', parent=self.styles['Normal'], fontName='Helvetica-Bold', alignment=TA_CENTER, textColor=colors.whitesmoke))
        self.styles.add(ParagraphStyle(name='Footer', parent=self.styles['Normal'], alignment=TA_CENTER, fontSize=8))
        self.styles.add(ParagraphStyle(name='NormalRight', parent=self.styles['Normal'], alignment=TA_RIGHT))
        self.styles.add(ParagraphStyle(name='NormalCenter', parent=self.styles['Normal'], alignment=TA_CENTER))
        self.styles.add(ParagraphStyle(name='GLAccountHeader', parent=self.styles['h3'], fontSize=10, spaceBefore=0.1*inch, spaceAfter=0.05*inch, alignment=TA_LEFT))

    async def export_report(self, report_data: Dict[str, Any], format_type: Literal["pdf", "excel", "gst_excel_detail"]) -> Optional[bytes]:
        title = report_data.get('title', "Financial Report")
        is_gst_detail_export = isinstance(report_data, GSTReturnData) and "gst_excel_detail" in format_type

        if format_type == "pdf":
            if title == "Balance Sheet": return await self._export_balance_sheet_to_pdf(report_data)
            elif title == "Profit & Loss Statement": return await self._export_profit_loss_to_pdf(report_data)
            elif title == "Trial Balance": return await self._export_trial_balance_to_pdf(report_data)
            elif title == "General Ledger": return await self._export_general_ledger_to_pdf(report_data)
            elif title == "Income Tax Computation": return await self._export_tax_computation_to_pdf(report_data)
            elif title == "Statement of Cash Flows": return await self._export_cash_flow_to_pdf(report_data)
            else: return self._export_generic_table_to_pdf(report_data) 
        elif format_type == "excel":
            if title == "Balance Sheet": return await self._export_balance_sheet_to_excel(report_data)
            elif title == "Profit & Loss Statement": return await self._export_profit_loss_to_excel(report_data)
            elif title == "Trial Balance": return await self._export_trial_balance_to_excel(report_data)
            elif title == "General Ledger": return await self._export_general_ledger_to_excel(report_data)
            elif title == "Income Tax Computation": return await self._export_tax_computation_to_excel(report_data)
            elif title == "Statement of Cash Flows": return await self._export_cash_flow_to_excel(report_data)
            else: return self._export_generic_table_to_excel(report_data) 
        elif is_gst_detail_export:
            return await self._export_gst_f5_details_to_excel(cast(GSTReturnData, report_data))
        else:
            self.app_core.logger.error(f"Unsupported report format or data type mismatch: {format_type}, type: {type(report_data)}")
            return None

    def _format_decimal(self, value: Optional[Decimal], places: int = 2, show_blank_for_zero: bool = False) -> str:
        if value is None: return "" if show_blank_for_zero else self._format_decimal(Decimal(0), places) 
        if not isinstance(value, Decimal): 
            try: value = Decimal(str(value))
            except InvalidOperation: return "ERR_DEC" 
        if show_blank_for_zero and value.is_zero(): return ""
        return f"{value:,.{places}f}"

    async def _get_company_name(self) -> str:
        settings = await self.app_core.company_settings_service.get_company_settings()
        return settings.company_name if settings else "Your Company"

    def _add_pdf_header_footer(self, canvas, doc, company_name: str, report_title: str, date_desc: str):
        canvas.saveState(); page_width = doc.width + doc.leftMargin + doc.rightMargin; header_y_start = doc.height + doc.topMargin - 0.5*inch
        canvas.setFont('Helvetica-Bold', 14); canvas.drawCentredString(page_width/2, header_y_start, company_name)
        canvas.setFont('Helvetica-Bold', 12); canvas.drawCentredString(page_width/2, header_y_start - 0.25*inch, report_title)
        canvas.setFont('Helvetica', 10); canvas.drawCentredString(page_width/2, header_y_start - 0.5*inch, date_desc)
        footer_y = 0.5*inch; canvas.setFont('Helvetica', 8); canvas.drawString(doc.leftMargin, footer_y, f"Generated on: {date.today().strftime('%d %b %Y')}"); canvas.drawRightString(page_width - doc.rightMargin, footer_y, f"Page {doc.page}"); canvas.restoreState()

    async def _export_balance_sheet_to_pdf(self, report_data: Dict[str, Any]) -> bytes:
        buffer = BytesIO(); company_name = await self._get_company_name(); report_title = report_data.get('title', "Balance Sheet"); date_desc = report_data.get('report_date_description', "")
        doc = SimpleDocTemplate(buffer, pagesize=A4,rightMargin=0.75*inch, leftMargin=0.75*inch,topMargin=1.25*inch, bottomMargin=0.75*inch)
        story: List[Any] = []; has_comparative = bool(report_data.get('comparative_date')); col_widths = [3.5*inch, 1.5*inch]; 
        if has_comparative: col_widths.append(1.5*inch)
        header_texts = ["Description", "Current Period"]; 
        if has_comparative: header_texts.append("Comparative")
        table_header_row = [Paragraph(text, self.styles['TableHeader']) for text in header_texts]
        def build_section_data(section_key: str, title: str):
            data: List[List[Any]] = []; section = report_data.get(section_key, {})
            data.append([Paragraph(title, self.styles['SectionHeader'])] + (["", ""] if has_comparative else [""]))
            for acc in section.get('accounts', []):
                row = [Paragraph(f"  {acc['code']} - {acc['name']}", self.styles['AccountName']), Paragraph(self._format_decimal(acc['balance']), self.styles['Amount'])]
                if has_comparative:
                    comp_bal_str = ""; 
                    if section.get('comparative_accounts'):
                        comp_acc_list = section['comparative_accounts']; comp_acc_match = next((ca for ca in comp_acc_list if ca['id'] == acc['id']), None)
                        if comp_acc_match and comp_acc_match.get('balance') is not None: comp_bal_str = self._format_decimal(comp_acc_match['balance'])
                    row.append(Paragraph(comp_bal_str, self.styles['Amount']))
                data.append(row)
            total_row = [Paragraph(f"Total {title}", self.styles['AccountNameBold']), Paragraph(self._format_decimal(section.get('total')), self.styles['AmountBold'])]
            if has_comparative: total_row.append(Paragraph(self._format_decimal(section.get('comparative_total')), self.styles['AmountBold']))
            data.append(total_row); data.append([Spacer(1, 0.1*inch)] * len(col_widths)); return data
        bs_data: List[List[Any]] = [table_header_row]; bs_data.extend(build_section_data('assets', 'Assets')); bs_data.extend(build_section_data('liabilities', 'Liabilities')); bs_data.extend(build_section_data('equity', 'Equity'))
        total_lia_eq_row = [Paragraph("Total Liabilities & Equity", self.styles['AccountNameBold']), Paragraph(self._format_decimal(report_data.get('total_liabilities_equity')), self.styles['AmountBold'])]
        if has_comparative: total_lia_eq_row.append(Paragraph(self._format_decimal(report_data.get('comparative_total_liabilities_equity')), self.styles['AmountBold']))
        bs_data.append(total_lia_eq_row); bs_table = Table(bs_data, colWidths=col_widths)
        style = TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4F81BD")), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('LINEABOVE', (0, -1), (-1,-1), 1, colors.black), ('LINEBELOW', (1, -1), (-1, -1), 1, colors.black, None, (2,2), 0, 1), ('LINEBELOW', (1, -1), (-1, -1), 0.5, colors.black, None, (1,1), 0, 1),])
        current_row_idx = 1 
        for section_key in ['assets', 'liabilities', 'equity']:
            if section_key in report_data: style.add('SPAN', (0, current_row_idx), (len(col_widths)-1, current_row_idx)); current_row_idx += len(report_data[section_key].get('accounts', [])) + 2 
        bs_table.setStyle(style); story.append(bs_table)
        if report_data.get('is_balanced') is False: story.append(Spacer(1, 0.2*inch)); story.append(Paragraph("Warning: Balance Sheet is out of balance!", ParagraphStyle(name='Warning', parent=self.styles['Normal'], textColor=colors.red, fontName='Helvetica-Bold')))
        doc.build(story, onFirstPage=lambda c, d: self._add_pdf_header_footer(c, d, company_name, report_title, date_desc), onLaterPages=lambda c, d: self._add_pdf_header_footer(c, d, company_name, report_title, date_desc)); return buffer.getvalue()

    async def _export_profit_loss_to_pdf(self, report_data: Dict[str, Any]) -> bytes:
        buffer = BytesIO(); company_name = await self._get_company_name(); report_title = report_data.get('title', "Profit & Loss Statement"); date_desc = report_data.get('report_date_description', "")
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=0.75*inch, leftMargin=0.75*inch,topMargin=1.25*inch, bottomMargin=0.75*inch)
        story: List[Any] = []; has_comparative = bool(report_data.get('comparative_start')); comp_header_text = "Comparative"; 
        if has_comparative and report_data.get('comparative_start') and report_data.get('comparative_end'): comp_start_str = report_data['comparative_start'].strftime('%d/%m/%y'); comp_end_str = report_data['comparative_end'].strftime('%d/%m/%y'); comp_header_text = f"Comp. ({comp_start_str}-{comp_end_str})"
        headers = ["Description", "Amount"]; 
        if has_comparative: headers.append(comp_header_text); 
        table_header_row = [Paragraph(text, self.styles['TableHeader']) for text in headers]; col_widths = [3.5*inch, 1.5*inch]; 
        if has_comparative: col_widths.append(1.5*inch)
        def build_pl_section_data(section_key: str, title: str, is_subtraction: bool = False):
            data: List[List[Any]] = []; section = report_data.get(section_key, {})
            data.append([Paragraph(title, self.styles['SectionHeader'])] + (["", ""] if has_comparative else [""]))
            for acc in section.get('accounts', []):
                balance = acc['balance']; row = [Paragraph(f"  {acc['code']} - {acc['name']}", self.styles['AccountName']), Paragraph(self._format_decimal(balance), self.styles['Amount'])]
                if has_comparative:
                    comp_bal_str = ""; 
                    if section.get('comparative_accounts'):
                        comp_acc_list = section['comparative_accounts']; comp_acc_match = next((ca for ca in comp_acc_list if ca['id'] == acc['id']), None)
                        if comp_acc_match and comp_acc_match.get('balance') is not None: comp_bal_str = self._format_decimal(comp_acc_match['balance'])
                    row.append(Paragraph(comp_bal_str, self.styles['Amount']))
                data.append(row)
            total = section.get('total'); total_row = [Paragraph(f"Total {title}", self.styles['AccountNameBold']), Paragraph(self._format_decimal(total), self.styles['AmountBold'])]
            if has_comparative: comp_total = section.get('comparative_total'); total_row.append(Paragraph(self._format_decimal(comp_total), self.styles['AmountBold']))
            data.append(total_row); data.append([Spacer(1, 0.1*inch)] * len(col_widths)); return data, total, section.get('comparative_total') if has_comparative else None
        pl_data: List[List[Any]] = [table_header_row]; rev_data, total_revenue, comp_total_revenue = build_pl_section_data('revenue', 'Revenue'); pl_data.extend(rev_data)
        exp_data, total_expenses, comp_total_expenses = build_pl_section_data('expenses', 'Operating Expenses', is_subtraction=True); pl_data.extend(exp_data)
        net_profit = report_data.get('net_profit', Decimal(0)); comp_net_profit = report_data.get('comparative_net_profit')
        net_profit_row = [Paragraph("Net Profit / (Loss)", self.styles['AccountNameBold']), Paragraph(self._format_decimal(net_profit), self.styles['AmountBold'])]
        if has_comparative: net_profit_row.append(Paragraph(self._format_decimal(comp_net_profit), self.styles['AmountBold']))
        pl_data.append(net_profit_row); pl_table = Table(pl_data, colWidths=col_widths)
        style = TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4F81BD")), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('LINEABOVE', (1, -1), (-1,-1), 1, colors.black), ('LINEBELOW', (1, -1), (-1, -1), 1, colors.black, None, (2,2), 0, 1), ('LINEBELOW', (1, -1), (-1, -1), 0.5, colors.black, None, (1,1), 0, 1),])
        current_row_idx = 1
        for section_key in ['revenue', 'expenses']: 
            if section_key in report_data: style.add('SPAN', (0, current_row_idx), (len(col_widths)-1, current_row_idx)); current_row_idx += len(report_data[section_key].get('accounts', [])) + 2
        pl_table.setStyle(style); story.append(pl_table)
        doc.build(story, onFirstPage=lambda c, d: self._add_pdf_header_footer(c, d, company_name, report_title, date_desc), onLaterPages=lambda c, d: self._add_pdf_header_footer(c, d, company_name, report_title, date_desc)); return buffer.getvalue()

    async def _export_trial_balance_to_pdf(self, report_data: Dict[str, Any]) -> bytes:
        buffer = BytesIO(); company_name = await self._get_company_name(); report_title = report_data.get('title', "Trial Balance"); date_desc = report_data.get('report_date_description', f"As of {date.today().strftime('%d %b %Y')}")
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=0.75*inch, leftMargin=0.75*inch,topMargin=1.25*inch, bottomMargin=0.75*inch)
        story: List[Any] = []; col_widths = [1.5*inch, 3*inch, 1.25*inch, 1.25*inch]; header_texts = ["Account Code", "Account Name", "Debit", "Credit"]
        table_header_row = [Paragraph(text, self.styles['TableHeader']) for text in header_texts]; tb_data: List[List[Any]] = [table_header_row]
        for acc in report_data.get('debit_accounts', []): tb_data.append([Paragraph(acc['code'], self.styles['Normal']), Paragraph(acc['name'], self.styles['Normal']), Paragraph(self._format_decimal(acc['balance']), self.styles['NormalRight']), Paragraph("", self.styles['NormalRight'])])
        for acc in report_data.get('credit_accounts', []): tb_data.append([Paragraph(acc['code'], self.styles['Normal']), Paragraph(acc['name'], self.styles['Normal']), Paragraph("", self.styles['NormalRight']), Paragraph(self._format_decimal(acc['balance']), self.styles['NormalRight'])])
        tb_data.append([Paragraph(""), Paragraph("TOTALS", self.styles['AccountNameBold']), Paragraph(self._format_decimal(report_data.get('total_debits')), self.styles['AmountBold']), Paragraph(self._format_decimal(report_data.get('total_credits')), self.styles['AmountBold'])])
        tb_table = Table(tb_data, colWidths=col_widths)
        style = TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4F81BD")), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('LINEABOVE', (2, -1), (3, -1), 1, colors.black), ('LINEBELOW', (2, -1), (3, -1), 1, colors.black), ('ALIGN', (2,1), (3,-1), 'RIGHT'),])
        tb_table.setStyle(style); story.append(tb_table)
        if report_data.get('is_balanced') is False: story.append(Spacer(1,0.2*inch)); story.append(Paragraph("Warning: Trial Balance is out of balance!", ParagraphStyle(name='Warning', parent=self.styles['Normal'], textColor=colors.red, fontName='Helvetica-Bold')))
        doc.build(story, onFirstPage=lambda c, d: self._add_pdf_header_footer(c, d, company_name, report_title, date_desc), onLaterPages=lambda c, d: self._add_pdf_header_footer(c, d, company_name, report_title, date_desc)); return buffer.getvalue()

    async def _export_general_ledger_to_pdf(self, report_data: Dict[str, Any]) -> bytes:
        buffer = BytesIO(); company_name = await self._get_company_name(); report_title = report_data.get('title', "General Ledger"); date_desc = report_data.get('report_date_description', "") 
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=0.5*inch, leftMargin=0.5*inch, topMargin=1.25*inch, bottomMargin=0.75*inch)
        story: List[Any] = []; story.append(Paragraph(f"Account: {report_data.get('account_code')} - {report_data.get('account_name')}", self.styles['GLAccountHeader'])); story.append(Spacer(1, 0.1*inch)); story.append(Paragraph(f"Opening Balance: {self._format_decimal(report_data.get('opening_balance'))}", self.styles['AccountNameBold'])); story.append(Spacer(1, 0.2*inch))
        headers = ["Date", "Entry No.", "JE Description", "Line Description", "Debit", "Credit", "Balance"]
        table_header_row = [Paragraph(h, self.styles['TableHeader']) for h in headers]; gl_data: List[List[Any]] = [table_header_row]
        for txn in report_data.get('transactions', []): gl_data.append([Paragraph(txn['date'].strftime('%d/%m/%Y') if isinstance(txn['date'], date) else str(txn['date']), self.styles['NormalCenter']), Paragraph(txn['entry_no'], self.styles['Normal']), Paragraph(txn.get('je_description','')[:40], self.styles['Normal']), Paragraph(txn.get('line_description','')[:40], self.styles['Normal']), Paragraph(self._format_decimal(txn['debit'], show_blank_for_zero=True), self.styles['NormalRight']), Paragraph(self._format_decimal(txn['credit'], show_blank_for_zero=True), self.styles['NormalRight']), Paragraph(self._format_decimal(txn['balance']), self.styles['NormalRight'])])
        col_widths_gl = [0.9*inch, 1.0*inch, 2.8*inch, 2.8*inch, 1.0*inch, 1.0*inch, 1.19*inch]
        gl_table = Table(gl_data, colWidths=col_widths_gl); style = TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4F81BD")), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('ALIGN', (0,1), (1,-1), 'CENTER'), ('ALIGN', (4,1), (6,-1), 'RIGHT'),]); gl_table.setStyle(style); story.append(gl_table); story.append(Spacer(1, 0.1*inch)); story.append(Paragraph(f"Closing Balance: {self._format_decimal(report_data.get('closing_balance'))}", self.styles['AmountBold']))
        doc.build(story, onFirstPage=lambda c, d: self._add_pdf_header_footer(c, d, company_name, report_title, date_desc), onLaterPages=lambda c, d: self._add_pdf_header_footer(c, d, company_name, report_title, date_desc)); return buffer.getvalue()

    async def _export_tax_computation_to_pdf(self, report_data: Dict[str, Any]) -> bytes:
        buffer = BytesIO(); company_name = await self._get_company_name(); report_title = report_data.get('title', "Tax Computation"); date_desc = report_data.get('report_date_description', "")
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=0.75*inch, leftMargin=0.75*inch, topMargin=1.25*inch, bottomMargin=0.75*inch)
        story: List[Any] = []; col_widths = [5*inch, 1.5*inch]; table_header_row = [Paragraph("Description", self.styles['TableHeader']), Paragraph("Amount", self.styles['TableHeader'])]; data: List[List[Any]] = [table_header_row]
        data.append([Paragraph("Net Profit Before Tax", self.styles['AccountNameBold']), Paragraph(self._format_decimal(report_data.get('net_profit_before_tax')), self.styles['AmountBold'])])
        data.append([Paragraph("Add: Non-Deductible Expenses", self.styles['SectionHeader']), ""]); 
        for adj in report_data.get('add_back_adjustments', []): data.append([Paragraph(f"  {adj['name']}", self.styles['AccountName']), Paragraph(self._format_decimal(adj['amount']), self.styles['Amount'])])
        data.append([Paragraph("Less: Non-Taxable Income", self.styles['SectionHeader']), ""])
        for adj in report_data.get('less_adjustments', []): data.append([Paragraph(f"  {adj['name']}", self.styles['AccountName']), Paragraph(f"({self._format_decimal(adj['amount'])})", self.styles['Amount'])])
        data.append([Paragraph("Chargeable Income", self.styles['AccountNameBold']), Paragraph(self._format_decimal(report_data.get('chargeable_income')), self.styles['AmountBold'])])
        data.append([Paragraph(f"Estimated Tax @ {report_data.get('tax_rate', 0):.2f}%", self.styles['AccountNameBold']), Paragraph(self._format_decimal(report_data.get('estimated_tax')), self.styles['AmountBold'])])
        table = Table(data, colWidths=col_widths)
        style = TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4F81BD")), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('LINEABOVE', (1, -1), (1,-1), 1, colors.black), ('LINEBELOW', (1, -1), (1, -1), 1, colors.black, None, (2,2), 0, 1),])
        style.add('SPAN', (0, 2), (1, 2)); style.add('SPAN', (0, len(data)-5), (1, len(data)-5)); 
        table.setStyle(style); story.append(table)
        doc.build(story, onFirstPage=lambda c, d: self._add_pdf_header_footer(c, d, company_name, report_title, date_desc), onLaterPages=lambda c, d: self._add_pdf_header_footer(c, d, company_name, report_title, date_desc)); return buffer.getvalue()

    async def _export_cash_flow_to_pdf(self, report_data: Dict[str, Any]) -> bytes:
        buffer = BytesIO(); company_name = await self._get_company_name(); report_title = report_data.get('title', "Statement of Cash Flows"); date_desc = report_data.get('report_date_description', "")
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=0.75*inch, leftMargin=0.75*inch, topMargin=1.25*inch, bottomMargin=0.75*inch)
        story: List[Any] = []; col_widths = [5*inch, 1.5*inch]; data: List[List[Any]] = []
        def add_row(desc: str, amt: Optional[Decimal] = None, indent: int = 0, is_bold: bool = False, is_header: bool = False, underline: bool = False):
            style_name = 'SectionHeader' if is_header else ('AccountNameBold' if is_bold else 'AccountName')
            amount_style_name = 'AmountBold' if is_bold else 'Amount'
            desc_p = Paragraph(f"{'&nbsp;'*4*indent}{desc}", self.styles[style_name])
            amt_p = Paragraph(f"({self._format_decimal(abs(amt))})" if amt is not None and amt < 0 else self._format_decimal(amt, show_blank_for_zero=(amt is None)), self.styles[amount_style_name])
            data.append([desc_p, amt_p])
            if underline: data.append([Spacer(1, 0.02*inch), Paragraph('<underline offset="-2" width="100%"> </underline>', self.styles['Amount'])])

        add_row("Cash flows from operating activities", is_header=True)
        add_row("Net income", report_data.get('net_income'), indent=1)
        add_row("Adjustments to reconcile net income to net cash provided by operating activities:", indent=1, is_bold=True)
        for adj in report_data.get('cfo_adjustments', []): add_row(adj['name'], adj['amount'], indent=2)
        add_row("Net cash provided by (used in) operating activities", report_data.get('net_cfo'), indent=0, is_bold=True, underline=True)
        data.append([Spacer(1, 0.1*inch)]*2)
        add_row("Cash flows from investing activities", is_header=True)
        for adj in report_data.get('cfi_adjustments', []): add_row(adj['name'], adj['amount'], indent=2)
        add_row("Net cash provided by (used in) investing activities", report_data.get('net_cfi'), indent=0, is_bold=True, underline=True)
        data.append([Spacer(1, 0.1*inch)]*2)
        add_row("Cash flows from financing activities", is_header=True)
        for adj in report_data.get('cff_adjustments', []): add_row(adj['name'], adj['amount'], indent=2)
        add_row("Net cash provided by (used in) financing activities", report_data.get('net_cff'), indent=0, is_bold=True, underline=True)
        data.append([Spacer(1, 0.2*inch)]*2)
        add_row("Net increase (decrease) in cash", report_data.get('net_change_in_cash'), is_bold=True)
        add_row("Cash at beginning of period", report_data.get('cash_at_start'), is_underline=True)
        add_row("Cash at end of period", report_data.get('cash_at_end'), is_bold=True, underline=True)
        
        table = Table(data, colWidths=col_widths); table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE')])); story.append(table)
        doc.build(story, onFirstPage=lambda c, d: self._add_pdf_header_footer(c, d, company_name, report_title, date_desc), onLaterPages=lambda c, d: self._add_pdf_header_footer(c, d, company_name, report_title, date_desc)); return buffer.getvalue()

    def _export_generic_table_to_pdf(self, report_data: Dict[str, Any]) -> bytes:
        self.app_core.logger.warning(f"Using generic PDF export for report: {report_data.get('title')}")
        buffer = BytesIO(); doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=0.5*inch, leftMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story: List[Any] = [Paragraph(report_data.get('title', "Report"), self.styles['h1'])]; 
        if report_data.get('report_date_description'): story.append(Paragraph(report_data.get('report_date_description'), self.styles['h3']))
        story.append(Spacer(1, 0.2*inch)); data_to_display = report_data.get('data_rows', []); headers = report_data.get('headers', [])
        if headers and data_to_display:
            table_data: List[List[Any]] = [[Paragraph(str(h), self.styles['TableHeader']) for h in headers]]
            for row_dict in data_to_display: table_data.append([Paragraph(str(row_dict.get(h_key, '')), self.styles['Normal']) for h_key in headers]) 
            num_cols = len(headers); col_widths_generic = [doc.width/num_cols]*num_cols
            table = Table(table_data, colWidths=col_widths_generic); table.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4F81BD"))])); story.append(table)
        else: story.append(Paragraph("No data or headers provided for generic export.", self.styles['Normal']))
        doc.build(story); return buffer.getvalue()

    def _apply_excel_header_style(self, cell, bold=True, size=12, alignment='center', fill_color: Optional[str]=None):
        cell.font = Font(bold=bold, size=size, color="FFFFFF" if fill_color else "000000")
        if alignment == 'center': cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        elif alignment == 'right': cell.alignment = Alignment(horizontal='right', vertical='center')
        else: cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        if fill_color: cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    def _apply_excel_amount_style(self, cell, bold=False, underline: Optional[str]=None):
        cell.number_format = '#,##0.00;[Red](#,##0.00);"-"' 
        cell.alignment = Alignment(horizontal='right')
        if bold: cell.font = Font(bold=True)
        if underline: cell.border = Border(bottom=Side(style=underline))
        
    async def _export_balance_sheet_to_excel(self, report_data: Dict[str, Any]) -> bytes:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Balance Sheet"; company_name = await self._get_company_name(); has_comparative = bool(report_data.get('comparative_date')); row_num = 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3 if has_comparative else 2); self._apply_excel_header_style(ws.cell(row=row_num, column=1, value=company_name), size=14); row_num +=1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3 if has_comparative else 2); self._apply_excel_header_style(ws.cell(row=row_num, column=1, value=report_data.get('title')), size=12); row_num +=1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3 if has_comparative else 2); self._apply_excel_header_style(ws.cell(row=row_num, column=1, value=report_data.get('report_date_description')), size=10, bold=False); row_num += 2
        headers = ["Description", "Current Period"]; 
        if has_comparative: headers.append("Comparative")
        for col, header_text in enumerate(headers, 1): self._apply_excel_header_style(ws.cell(row=row_num, column=col, value=header_text), alignment='center' if col>0 else 'left', fill_color="4F81BD")
        row_num +=1
        def write_section(section_key: str, title: str):
            nonlocal row_num; section = report_data.get(section_key, {})
            ws.cell(row=row_num, column=1, value=title).font = Font(bold=True, size=11, color="2F5496"); row_num +=1
            for acc in section.get('accounts', []):
                ws.cell(row=row_num, column=1, value=f"  {acc['code']} - {acc['name']}")
                self._apply_excel_amount_style(ws.cell(row=row_num, column=2, value=float(acc['balance'] if acc.get('balance') is not None else 0)))
                if has_comparative:
                    comp_val = None
                    if section.get('comparative_accounts'):
                        comp_acc_match = next((ca for ca in section['comparative_accounts'] if ca['id'] == acc['id']), None)
                        if comp_acc_match and comp_acc_match.get('balance') is not None: comp_val = float(comp_acc_match['balance'])
                    self._apply_excel_amount_style(ws.cell(row=row_num, column=3, value=comp_val if comp_val is not None else "-"))
                row_num +=1
            ws.cell(row=row_num, column=1, value=f"Total {title}").font = Font(bold=True)
            self._apply_excel_amount_style(ws.cell(row=row_num, column=2, value=float(section.get('total',0))), bold=True, underline="thin")
            if has_comparative: self._apply_excel_amount_style(ws.cell(row=row_num, column=3, value=float(section.get('comparative_total',0))), bold=True, underline="thin")
            row_num += 2
        write_section('assets', 'Assets'); write_section('liabilities', 'Liabilities'); write_section('equity', 'Equity')
        ws.cell(row=row_num, column=1, value="Total Liabilities & Equity").font = Font(bold=True, size=11)
        self._apply_excel_amount_style(ws.cell(row=row_num, column=2, value=float(report_data.get('total_liabilities_equity',0))), bold=True, underline="double")
        if has_comparative: self._apply_excel_amount_style(ws.cell(row=row_num, column=3, value=float(report_data.get('comparative_total_liabilities_equity',0))), bold=True, underline="double")
        row_num +=1
        if report_data.get('is_balanced') is False: ws.cell(row=row_num, column=1, value="Warning: Balance Sheet out of balance!").font = Font(color="FF0000", bold=True)
        for i, col_letter in enumerate(['A','B','C'][:len(headers)]): ws.column_dimensions[col_letter].width = 45 if i==0 else 20
        excel_bytes_io = BytesIO(); wb.save(excel_bytes_io); return excel_bytes_io.getvalue()

    async def _export_profit_loss_to_excel(self, report_data: Dict[str, Any]) -> bytes:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Profit and Loss"; company_name = await self._get_company_name(); has_comparative = bool(report_data.get('comparative_start')); row_num = 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3 if has_comparative else 2); self._apply_excel_header_style(ws.cell(row=row_num, column=1, value=company_name), size=14); row_num +=1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3 if has_comparative else 2); self._apply_excel_header_style(ws.cell(row=row_num, column=1, value=report_data.get('title')), size=12); row_num +=1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3 if has_comparative else 2); self._apply_excel_header_style(ws.cell(row=row_num, column=1, value=report_data.get('report_date_description')), size=10, bold=False); row_num += 2
        headers = ["Description", "Current Period"]; 
        if has_comparative: headers.append("Comparative")
        for col, header_text in enumerate(headers, 1): self._apply_excel_header_style(ws.cell(row=row_num, column=col, value=header_text), alignment='center' if col>0 else 'left', fill_color="4F81BD")
        row_num +=1
        def write_pl_section(section_key: str, title: str):
            nonlocal row_num; section = report_data.get(section_key, {})
            ws.cell(row=row_num, column=1, value=title).font = Font(bold=True, size=11, color="2F5496"); row_num +=1
            for acc in section.get('accounts', []):
                ws.cell(row=row_num, column=1, value=f"  {acc['code']} - {acc['name']}")
                self._apply_excel_amount_style(ws.cell(row=row_num, column=2, value=float(acc['balance'] if acc.get('balance') is not None else 0)))
                if has_comparative:
                    comp_val = None
                    if section.get('comparative_accounts'):
                        comp_acc_match = next((ca for ca in section['comparative_accounts'] if ca['id'] == acc['id']), None)
                        if comp_acc_match and comp_acc_match.get('balance') is not None: comp_val = float(comp_acc_match['balance'])
                    self._apply_excel_amount_style(ws.cell(row=row_num, column=3, value=comp_val if comp_val is not None else "-"))
                row_num +=1
            ws.cell(row=row_num, column=1, value=f"Total {title}").font = Font(bold=True)
            self._apply_excel_amount_style(ws.cell(row=row_num, column=2, value=float(section.get('total',0))), bold=True, underline="thin")
            if has_comparative: self._apply_excel_amount_style(ws.cell(row=row_num, column=3, value=float(section.get('comparative_total',0))), bold=True, underline="thin")
            row_num +=1; return section.get('total', Decimal(0)), section.get('comparative_total') if has_comparative else Decimal(0)
        total_revenue, comp_total_revenue = write_pl_section('revenue', 'Revenue'); row_num +=1 
        total_expenses, comp_total_expenses = write_pl_section('expenses', 'Operating Expenses'); row_num +=1 
        ws.cell(row=row_num, column=1, value="Net Profit / (Loss)").font = Font(bold=True, size=11)
        self._apply_excel_amount_style(ws.cell(row=row_num, column=2, value=float(report_data.get('net_profit',0))), bold=True, underline="double")
        if has_comparative: self._apply_excel_amount_style(ws.cell(row=row_num, column=3, value=float(report_data.get('comparative_net_profit',0))), bold=True, underline="double")
        for i, col_letter in enumerate(['A','B','C'][:len(headers)]): ws.column_dimensions[col_letter].width = 45 if i==0 else 20
        excel_bytes_io = BytesIO(); wb.save(excel_bytes_io); return excel_bytes_io.getvalue()
    
    async def _export_trial_balance_to_excel(self, report_data: Dict[str, Any]) -> bytes:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Trial Balance"; company_name = await self._get_company_name(); row_num = 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4); self._apply_excel_header_style(ws.cell(row=row_num, column=1, value=company_name), size=14, alignment='center'); row_num += 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4); self._apply_excel_header_style(ws.cell(row=row_num, column=1, value=report_data.get('title')), size=12, alignment='center'); row_num += 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4); self._apply_excel_header_style(ws.cell(row=row_num, column=1, value=report_data.get('report_date_description')), size=10, bold=False, alignment='center'); row_num += 2
        headers = ["Account Code", "Account Name", "Debit", "Credit"]
        for col, header_text in enumerate(headers, 1): self._apply_excel_header_style(ws.cell(row=row_num, column=col, value=header_text), fill_color="4F81BD", alignment='center' if col > 1 else 'left')
        row_num += 1
        for acc in report_data.get('debit_accounts', []): ws.cell(row=row_num, column=1, value=acc['code']); ws.cell(row=row_num, column=2, value=acc['name']); self._apply_excel_amount_style(ws.cell(row=row_num, column=3, value=float(acc['balance'] if acc.get('balance') is not None else 0))); ws.cell(row=row_num, column=4, value=None); row_num += 1
        for acc in report_data.get('credit_accounts', []): ws.cell(row=row_num, column=1, value=acc['code']); ws.cell(row=row_num, column=2, value=acc['name']); ws.cell(row=row_num, column=3, value=None); self._apply_excel_amount_style(ws.cell(row=row_num, column=4, value=float(acc['balance'] if acc.get('balance') is not None else 0))); row_num += 1
        row_num +=1; ws.cell(row=row_num, column=2, value="TOTALS").font = Font(bold=True); ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='right')
        self._apply_excel_amount_style(ws.cell(row=row_num, column=3, value=float(report_data.get('total_debits', 0))), bold=True, underline="thin"); self._apply_excel_amount_style(ws.cell(row=row_num, column=4, value=float(report_data.get('total_credits', 0))), bold=True, underline="thin")
        if report_data.get('is_balanced', True) is False: row_num +=2; ws.cell(row=row_num, column=1, value="Warning: Trial Balance is out of balance!").font = Font(color="FF0000", bold=True)
        ws.column_dimensions[get_column_letter(1)].width = 15; ws.column_dimensions[get_column_letter(2)].width = 45; ws.column_dimensions[get_column_letter(3)].width = 20; ws.column_dimensions[get_column_letter(4)].width = 20
        excel_bytes_io = BytesIO(); wb.save(excel_bytes_io); return excel_bytes_io.getvalue()

    async def _export_general_ledger_to_excel(self, report_data: Dict[str, Any]) -> bytes:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = f"GL-{report_data.get('account_code', 'Account')}"[:30]; company_name = await self._get_company_name(); row_num = 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7); self._apply_excel_header_style(ws.cell(row=row_num, column=1, value=company_name), size=14, alignment='center'); row_num += 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7); self._apply_excel_header_style(ws.cell(row=row_num, column=1, value=report_data.get('title')), size=12, alignment='center'); row_num += 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7); self._apply_excel_header_style(ws.cell(row=row_num, column=1, value=report_data.get('report_date_description')), size=10, bold=False, alignment='center'); row_num += 2
        headers = ["Date", "Entry No.", "JE Description", "Line Description", "Debit", "Credit", "Balance"]
        for col, header_text in enumerate(headers, 1): self._apply_excel_header_style(ws.cell(row=row_num, column=col, value=header_text), fill_color="4F81BD", alignment='center' if col > 3 else 'left')
        row_num += 1
        ws.cell(row=row_num, column=4, value="Opening Balance").font = Font(bold=True); ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='right'); self._apply_excel_amount_style(ws.cell(row=row_num, column=7, value=float(report_data.get('opening_balance',0))), bold=True); row_num += 1
        for txn in report_data.get('transactions', []):
            ws.cell(row=row_num, column=1, value=txn['date'].strftime('%d/%m/%Y') if isinstance(txn['date'], date) else txn['date']); ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=2, value=txn['entry_no']); ws.cell(row=row_num, column=3, value=txn.get('je_description','')); ws.cell(row=row_num, column=4, value=txn.get('line_description',''))
            self._apply_excel_amount_style(ws.cell(row=row_num, column=5, value=float(txn['debit'] if txn['debit'] else 0))); self._apply_excel_amount_style(ws.cell(row=row_num, column=6, value=float(txn['credit'] if txn['credit'] else 0))); self._apply_excel_amount_style(ws.cell(row=row_num, column=7, value=float(txn['balance']))); row_num += 1
        ws.cell(row=row_num, column=4, value="Closing Balance").font = Font(bold=True); ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='right'); self._apply_excel_amount_style(ws.cell(row=row_num, column=7, value=float(report_data.get('closing_balance',0))), bold=True, underline="thin")
        ws.column_dimensions[get_column_letter(1)].width = 12; ws.column_dimensions[get_column_letter(2)].width = 15; ws.column_dimensions[get_column_letter(3)].width = 40; ws.column_dimensions[get_column_letter(4)].width = 40
        for i in [5,6,7]: ws.column_dimensions[get_column_letter(i)].width = 18
        excel_bytes_io = BytesIO(); wb.save(excel_bytes_io); return excel_bytes_io.getvalue()

    async def _export_tax_computation_to_excel(self, report_data: Dict[str, Any]) -> bytes:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Tax Computation"; company_name = await self._get_company_name(); row_num = 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2); self._apply_excel_header_style(ws.cell(row=row_num, column=1, value=company_name), size=14, alignment='center'); row_num += 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2); self._apply_excel_header_style(ws.cell(row=row_num, column=1, value=report_data.get('title')), size=12, alignment='center'); row_num += 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2); self._apply_excel_header_style(ws.cell(row=row_num, column=1, value=report_data.get('report_date_description')), size=10, bold=False, alignment='center'); row_num += 2
        for col, header_text in enumerate(["Description", "Amount"], 1): self._apply_excel_header_style(ws.cell(row=row_num, column=col, value=header_text), fill_color="4F81BD")
        row_num +=1
        ws.cell(row=row_num, column=1, value="Net Profit Before Tax").font = Font(bold=True); self._apply_excel_amount_style(ws.cell(row=row_num, column=2, value=float(report_data.get('net_profit_before_tax',0))), bold=True); row_num += 2
        ws.cell(row=row_num, column=1, value="Add: Non-Deductible Expenses").font = Font(bold=True, color="2F5496"); row_num +=1
        for adj in report_data.get('add_back_adjustments', []): ws.cell(row=row_num, column=1, value=f"  {adj['name']}"); self._apply_excel_amount_style(ws.cell(row=row_num, column=2, value=float(adj['amount']))); row_num += 1
        ws.cell(row=row_num, column=1, value="Less: Non-Taxable Income").font = Font(bold=True, color="2F5496"); row_num +=1
        for adj in report_data.get('less_adjustments', []): ws.cell(row=row_num, column=1, value=f"  {adj['name']}"); self._apply_excel_amount_style(ws.cell(row=row_num, column=2, value=float(adj['amount']))); row_num += 1
        row_num += 1; ws.cell(row=row_num, column=1, value="Chargeable Income").font = Font(bold=True); self._apply_excel_amount_style(ws.cell(row=row_num, column=2, value=float(report_data.get('chargeable_income',0))), bold=True, underline="thin"); row_num += 2
        ws.cell(row=row_num, column=1, value=f"Estimated Tax @ {report_data.get('tax_rate', 0):.2f}%").font = Font(bold=True); self._apply_excel_amount_style(ws.cell(row=row_num, column=2, value=float(report_data.get('estimated_tax',0))), bold=True, underline="double");
        ws.column_dimensions[get_column_letter(1)].width = 60; ws.column_dimensions[get_column_letter(2)].width = 20
        excel_bytes_io = BytesIO(); wb.save(excel_bytes_io); return excel_bytes_io.getvalue()

    async def _export_cash_flow_to_excel(self, report_data: Dict[str, Any]) -> bytes:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Cash Flow Statement"; company_name = await self._get_company_name(); row_num = 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2); self._apply_excel_header_style(ws.cell(row=row_num, column=1, value=company_name), size=14); row_num += 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2); self._apply_excel_header_style(ws.cell(row=row_num, column=1, value=report_data.get('title')), size=12); row_num += 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2); self._apply_excel_header_style(ws.cell(row=row_num, column=1, value=report_data.get('report_date_description')), size=10, bold=False); row_num += 2
        def write_row(desc: str, amt: Optional[Decimal], indent: int=0, is_bold: bool=False, is_header: bool=False, underline: Optional[str]=None):
            nonlocal row_num; cell1 = ws.cell(row=row_num, column=1, value=f"{' ' * (indent * 2)}{desc}"); cell2 = ws.cell(row=row_num, column=2, value=float(amt) if amt is not None else None); self._apply_excel_amount_style(cell2, bold=is_bold, underline=underline); cell1.font = Font(bold=is_bold, color="2F5496" if is_header else "000000"); cell1.alignment = Alignment(indent=indent); row_num += 1
        
        write_row("Cash flows from operating activities", is_header=True)
        write_row("Net income", report_data.get('net_income'), indent=1); write_row("Adjustments:", indent=1, is_bold=True)
        for adj in report_data.get('cfo_adjustments',[]): write_row(adj['name'], adj['amount'], indent=2)
        write_row("Net cash provided by (used in) operating activities", report_data.get('net_cfo'), indent=0, is_bold=True, underline="thin"); row_num += 1
        write_row("Cash flows from investing activities", is_header=True)
        for adj in report_data.get('cfi_adjustments',[]): write_row(adj['name'], adj['amount'], indent=2)
        write_row("Net cash provided by (used in) investing activities", report_data.get('net_cfi'), indent=0, is_bold=True, underline="thin"); row_num += 1
        write_row("Cash flows from financing activities", is_header=True)
        for adj in report_data.get('cff_adjustments',[]): write_row(adj['name'], adj['amount'], indent=2)
        write_row("Net cash provided by (used in) financing activities", report_data.get('net_cff'), indent=0, is_bold=True, underline="thin"); row_num += 2
        write_row("Net increase (decrease) in cash", report_data.get('net_change_in_cash'), is_bold=True); write_row("Cash at beginning of period", report_data.get('cash_at_start'), underline="thin"); write_row("Cash at end of period", report_data.get('cash_at_end'), is_bold=True, underline="double")
        ws.column_dimensions['A'].width = 60; ws.column_dimensions['B'].width = 20
        excel_bytes_io = BytesIO(); wb.save(excel_bytes_io); return excel_bytes_io.getvalue()
        
    async def _export_gst_f5_details_to_excel(self, report_data: GSTReturnData) -> bytes:
        wb = openpyxl.Workbook()
        ws_summary = wb.active; ws_summary.title = "GST F5 Summary"; company_name = await self._get_company_name(); row_num = 1
        ws_summary.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3); self._apply_excel_header_style(ws_summary.cell(row=row_num, column=1, value=company_name), size=14); row_num += 1
        ws_summary.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3); self._apply_excel_header_style(ws_summary.cell(row=row_num, column=1, value="GST F5 Return"), size=12); row_num += 1
        ws_summary.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3); date_desc = f"For period: {report_data.start_date.strftime('%d %b %Y')} to {report_data.end_date.strftime('%d %b %Y')}"; self._apply_excel_header_style(ws_summary.cell(row=row_num, column=1, value=date_desc), size=10, bold=False); row_num += 2
        f5_boxes = [("1. Standard-Rated Supplies", report_data.standard_rated_supplies),("2. Zero-Rated Supplies", report_data.zero_rated_supplies),("3. Exempt Supplies", report_data.exempt_supplies),("4. Total Supplies (1+2+3)", report_data.total_supplies, True),("5. Taxable Purchases", report_data.taxable_purchases),("6. Output Tax Due", report_data.output_tax),("7. Input Tax and Refunds Claimed", report_data.input_tax),("8. GST Adjustments (e.g. Bad Debt Relief)", report_data.tax_adjustments),("9. Net GST Payable / (Claimable)", report_data.tax_payable, True)]
        for desc, val, *is_bold in f5_boxes: ws_summary.cell(row=row_num, column=1, value=desc).font = Font(bold=bool(is_bold and is_bold[0])); self._apply_excel_amount_style(ws_summary.cell(row=row_num, column=2, value=float(val)), bold=bool(is_bold and is_bold[0])); row_num +=1
        ws_summary.column_dimensions['A'].width = 45; ws_summary.column_dimensions['B'].width = 20
        detail_headers = ["Date", "Doc No.", "Entity", "Description", "GL Code", "GL Name", "Net Amount", "GST Amount", "Tax Code"]
        box_map_to_detail_key = {"Box 1: Std Supplies": "box1_standard_rated_supplies","Box 2: Zero Supplies": "box2_zero_rated_supplies","Box 3: Exempt Supplies": "box3_exempt_supplies","Box 5: Taxable Purchases": "box5_taxable_purchases","Box 6: Output Tax": "box6_output_tax_details","Box 7: Input Tax": "box7_input_tax_details"}
        if report_data.detailed_breakdown:
            for sheet_title_prefix, detail_key in box_map_to_detail_key.items():
                transactions: List[GSTTransactionLineDetail] = report_data.detailed_breakdown.get(detail_key, [])
                if not transactions: continue
                ws_detail = wb.create_sheet(title=sheet_title_prefix[:30]); row_num_detail = 1
                for col, header_text in enumerate(detail_headers, 1): self._apply_excel_header_style(ws_detail.cell(row=row_num_detail, column=col, value=header_text), fill_color="4F81BD")
                row_num_detail +=1
                for txn in transactions:
                    ws_detail.cell(row=row_num_detail, column=1, value=txn.transaction_date.strftime('%d/%m/%Y') if txn.transaction_date else None); ws_detail.cell(row=row_num_detail, column=2, value=txn.document_no); ws_detail.cell(row=row_num_detail, column=3, value=txn.entity_name); ws_detail.cell(row=row_num_detail, column=4, value=txn.description); ws_detail.cell(row=row_num_detail, column=5, value=txn.account_code); ws_detail.cell(row=row_num_detail, column=6, value=txn.account_name); self._apply_excel_amount_style(ws_detail.cell(row=row_num_detail, column=7, value=float(txn.net_amount))); self._apply_excel_amount_style(ws_detail.cell(row=row_num_detail, column=8, value=float(txn.gst_amount))); ws_detail.cell(row=row_num_detail, column=9, value=txn.tax_code_applied); row_num_detail += 1
                for i, col_letter in enumerate([get_column_letter(j+1) for j in range(len(detail_headers))]):
                    if i in [0,1,8]: ws_detail.column_dimensions[col_letter].width = 15
                    elif i in [2,3,4,5]: ws_detail.column_dimensions[col_letter].width = 30
                    else: ws_detail.column_dimensions[col_letter].width = 18
        excel_bytes_io = BytesIO(); wb.save(excel_bytes_io); return excel_bytes_io.getvalue()
        
    def _export_generic_table_to_excel(self, report_data: Dict[str, Any]) -> bytes:
        self.app_core.logger.warning(f"Using generic Excel export for report: {report_data.get('title')}")
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = report_data.get('title', "Report")[:30] # type: ignore
        row_num = 1; ws.cell(row=row_num, column=1, value=report_data.get('title')).font = Font(bold=True, size=14); row_num += 1 # type: ignore
        if report_data.get('report_date_description'): ws.cell(row=row_num, column=1, value=report_data.get('report_date_description')).font = Font(italic=True); row_num += 1 # type: ignore
        row_num += 1; headers = report_data.get('headers', []); data_rows = report_data.get('data_rows', []) 
        if headers:
            for col, header_text in enumerate(headers,1): self._apply_excel_header_style(ws.cell(row=row_num, column=col, value=header_text), fill_color="4F81BD")
            row_num +=1
        if data_rows:
            for data_item in data_rows:
                if isinstance(data_item, dict):
                    for col, header_text in enumerate(headers, 1): ws.cell(row=row_num, column=col, value=data_item.get(header_text, data_item.get(header_text.lower().replace(' ','_'))))
                elif isinstance(data_item, list):
                     for col, cell_val in enumerate(data_item, 1): ws.cell(row=row_num, column=col, value=cell_val)
                row_num +=1
        for col_idx_generic in range(1, len(headers) + 1): ws.column_dimensions[get_column_letter(col_idx_generic)].width = 20
        excel_bytes_io = BytesIO(); wb.save(excel_bytes_io); return excel_bytes_io.getvalue()
